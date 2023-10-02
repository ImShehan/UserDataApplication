from flask import Flask, render_template, request
from openpyxl import load_workbook, Workbook
from datetime import datetime

app = Flask(__name__)

# Function to create or load the Excel file
def get_workbook():
    try:
        workbook = load_workbook("user_data.xlsx")
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Name", "Email", "Employ Code", "Optical Message", "Date and Time"])
    return workbook

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/submit', methods=['POST'])
def submit():
    name = request.form.get('name')
    email = request.form.get('email')
    emp_code = request.form.get('empCode')
    optical_message = request.form.get('opticalMessage')
    
    # Add the current date and time
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    workbook = get_workbook()
    sheet = workbook.active

    # Add the submitted data along with the timestamp to the Excel file
    sheet.append([name, email, emp_code, optical_message, timestamp])

    # Save the Excel file
    excel_file_name = "user_data.xlsx"
    workbook.save(excel_file_name)

    return render_template('index.html', message="Data saved successfully!")

if __name__ == '__main__':
    app.run(debug=True)
